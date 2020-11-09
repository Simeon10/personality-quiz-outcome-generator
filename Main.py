from openpyxl import load_workbook


def getResultsList():  # Scrolls through the excel file to get all the names of all of the possible results
    resultsSheet = wb["Results"]
    for names in resultsSheet.iter_rows(min_row=1, max_row=1, min_col=2, values_only=True):
        for name in names:
            resultNames.append(name)


def getQuestionsList():  # Scrolls through excel file to get all the questions and answers and puts them into string arrays
    singleAnswers = []
    singleQuestions = []
    for sheet in wb:
        if not sheet.title == "Results":
            for answer in sheet.iter_rows(min_row=3, min_col=2, max_col=len(resultNames)+1, values_only=True):
                for x in answer:
                    singleAnswers.append(x)
                singleQuestions.append(singleAnswers)
                singleAnswers = []
            questionsAnswers.append(singleQuestions)
            singleQuestions = []
    questionsAnswers.sort(key=len, reverse=True)


def setVariables():  # Sets up the variables to proper length
    for x in range(numberOfQuestions):
        questionPointsPerRound.append(0)
    for x in range(len(resultNames)):
        finalResults.append(0)
        roundResults.append(0)


def scroll(question):  # Scrolls through all the possible outcomes
    global counter
    global firstTime
    for answer in question:
        questionPointsPerRound[counter] = answer
        if counter == len(questionsAnswers) - 1:  # If last question is reached, add up the scores
            addUpScores()
        else:  # Else, recurse
            if firstTime:
                firstTime = False
            else:
                counter += 1
            scroll(questionsAnswers[counter])  # Recursion occurs
    counter -= 1


def addUpScores():  # Adds up all the scores for each category for the particular round and picks a winning category
    localCounter = 0
    topScore = 0
    topScoreIndex = 0
    for answers in questionPointsPerRound:  # Adds up all the points for each category
        for x in answers:
            roundResults[localCounter] += x
            localCounter += 1
        localCounter = 0
    for x in range(len(roundResults)):  # Determines a winning category
        if roundResults[x] > topScore:
            topScoreIndex = x
            topScore = roundResults[x]
    finalResults[topScoreIndex] += 1  # Adds point to winning category
    for x in range(len(roundResults)):
        roundResults[x] = 0  # Resets variable


def finalScores():  # Prints out final result
    localCounter = 0
    totalPossibilities = sum(finalResults)
    for name in resultNames:
        possibilities = finalResults[localCounter]
        print(name + ": " + str(possibilities) + " (" + str(round(possibilities/totalPossibilities*100, 1)) + "%)")
        localCounter += 1
    print("Total number of possibilities = " + str(totalPossibilities))


resultNames = []
counter = 0
firstTime = True
questionsAnswers = []
finalResults = []
roundResults = []
questionPointsPerRound = []

wbName = input("Enter the full name of your excel file (excluding the \".xlsx\")")

wb = load_workbook(wbName + ".xlsx")
numberOfQuestions = len(wb.sheetnames) - 1

getResultsList()
getQuestionsList()
setVariables()
scroll([0])
finalScores()
